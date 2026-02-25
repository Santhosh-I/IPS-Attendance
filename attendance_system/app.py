from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file
import psycopg2
from psycopg2.extras import RealDictCursor
from datetime import datetime, date, time, timezone, timedelta
import os
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# IST timezone (UTC+5:30)
IST = timezone(timedelta(hours=5, minutes=30))

app = Flask(__name__)

DATABASE_URL = "postgresql://postgres.xuirqkdtrvkhjirrnmla:ipsattendance0000830245@aws-1-ap-south-1.pooler.supabase.com:6543/postgres"

@app.template_filter('to12hr')
def to12hr_filter(value):
    if not value:
        return '-'
    if isinstance(value, time):
        return value.strftime('%I:%M:%S %p')
    try:
        t = datetime.strptime(str(value), '%H:%M:%S')
        return t.strftime('%I:%M:%S %p')
    except ValueError:
        return str(value)

def _time_str(value):
    """Convert time/datetime.time to string for JSON responses."""
    if value is None:
        return '-'
    if isinstance(value, time):
        return value.strftime('%H:%M:%S')
    return str(value)

def _date_str(value):
    """Convert date/datetime.date to string for JSON responses."""
    if value is None:
        return ''
    if isinstance(value, date):
        return value.strftime('%Y-%m-%d')
    return str(value)

# -------------------------
# Database Connection
# -------------------------
def get_db():
    return psycopg2.connect(
        DATABASE_URL,
        cursor_factory=RealDictCursor
    )

# -------------------------
# Initialize Database
# -------------------------
def init_db():
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS students (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL,
            roll_number TEXT,
            rfid_uid TEXT UNIQUE NOT NULL
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS attendance (
            id SERIAL PRIMARY KEY,
            student_id INTEGER REFERENCES students(id),
            date DATE,
            in_time TIME,
            out_time TIME,
            entry_number INTEGER DEFAULT 1
        )
    ''')

    conn.commit()
    conn.close()

init_db()

# -------------------------
# Routes
# -------------------------
@app.route("/", methods=["GET", "POST"])
def tap():
    message = ""
    status = ""

    if request.method == "POST":
        uid = request.form["rfid"].strip()

        conn = get_db()
        cursor = conn.cursor()

        cursor.execute(
            "SELECT * FROM students WHERE rfid_uid = %s",
            (uid,)
        )
        student = cursor.fetchone()

        if not student:
            message = "ID Not Registered"
            status = "error"
        else:
            today = datetime.now(IST).strftime("%Y-%m-%d")
            current_time = datetime.now(IST).strftime("%H:%M:%S")

            cursor.execute(
                "SELECT * FROM attendance WHERE student_id = %s AND date = %s ORDER BY id DESC LIMIT 1",
                (student["id"], today),
            )
            latest = cursor.fetchone()

            if not latest or (latest["in_time"] and latest["out_time"]):
                entry_num = (latest["entry_number"] + 1) if latest else 1

                cursor.execute(
                    "INSERT INTO attendance (student_id, date, in_time, entry_number) VALUES (%s, %s, %s, %s)",
                    (student["id"], today, current_time, entry_num),
                )
                message = f"{student['name']} — IN Time Marked (Entry #{entry_num})"
                status = "in"
            else:
                cursor.execute(
                    "UPDATE attendance SET out_time = %s WHERE id = %s",
                    (current_time, latest["id"]),
                )
                message = f"{student['name']} — OUT Time Marked"
                status = "out"

            conn.commit()

        conn.close()

    return render_template("index.html", message=message, status=status)

@app.route("/dashboard")
def dashboard():
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('''
        SELECT students.name, attendance.date,
               attendance.entry_number, attendance.in_time, attendance.out_time
        FROM attendance
        JOIN students ON students.id = attendance.student_id
        ORDER BY attendance.date DESC, attendance.id DESC
    ''')
    records = cursor.fetchall()

    conn.close()
    return render_template("dashboard.html", records=records)

@app.route("/history")
def history():
    return render_template("history.html")

@app.route("/api/attendance/<date>")
def attendance_by_date(date):
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('''
        SELECT students.name, students.roll_number,
               attendance.in_time, attendance.out_time,
               attendance.entry_number
        FROM attendance
        JOIN students ON students.id = attendance.student_id
        WHERE attendance.date = %s
    ''', (date,))
    records = cursor.fetchall()

    conn.close()

    data = []
    for r in records:
        data.append({
            "name": r["name"],
            "roll_number": r["roll_number"] or "-",
            "in_time": _time_str(r["in_time"]),
            "out_time": _time_str(r["out_time"]),
            "entry_number": r["entry_number"]
        })

    return jsonify(data)

@app.route("/api/attendance-dates")
def attendance_dates():
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute(
        "SELECT DISTINCT date FROM attendance ORDER BY date"
    )
    dates = cursor.fetchall()

    conn.close()
    return jsonify([_date_str(d["date"]) for d in dates])

@app.route("/verify-admin", methods=["POST"])
def verify_admin():
    password = request.form.get("password", "")
    if password == "ips@2026":
        return redirect("/add-member")
    return redirect("/dashboard")

@app.route("/add-member", methods=["GET", "POST"])
def add_member():
    message = ""
    status = ""

    if request.method == "POST":
        name = request.form["name"].strip()
        roll_number = request.form.get("roll_number", "").strip()
        rfid_uid = request.form["rfid_uid"].strip()

        conn = get_db()
        cursor = conn.cursor()

        try:
            cursor.execute(
                "INSERT INTO students (name, roll_number, rfid_uid) VALUES (%s, %s, %s)",
                (name, roll_number, rfid_uid),
            )
            conn.commit()
            message = f"{name} added successfully!"
            status = "in"
        except psycopg2.IntegrityError:
            conn.rollback()
            message = "This RFID UID is already registered!"
            status = "error"
        finally:
            conn.close()

    return render_template("add_member.html", message=message, status=status)

@app.route("/members")
def members():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM students ORDER BY name")
    students = cursor.fetchall()
    conn.close()
    return render_template("members.html", students=students)

@app.route("/edit-member/<int:id>", methods=["GET", "POST"])
def edit_member(id):
    conn = get_db()
    cursor = conn.cursor()
    message = ""
    status = ""

    if request.method == "POST":
        name = request.form["name"].strip()
        roll_number = request.form.get("roll_number", "").strip()

        cursor.execute(
            "UPDATE students SET name = %s, roll_number = %s WHERE id = %s",
            (name, roll_number, id),
        )
        conn.commit()
        message = "Updated successfully!"
        status = "in"

    cursor.execute("SELECT * FROM students WHERE id = %s", (id,))
    student = cursor.fetchone()
    conn.close()

    if not student:
        return redirect("/members")

    return render_template("edit_member.html", student=student, message=message, status=status)

@app.route("/delete-member/<int:id>", methods=["POST"])
def delete_member(id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM attendance WHERE student_id = %s", (id,))
    cursor.execute("DELETE FROM students WHERE id = %s", (id,))
    conn.commit()
    conn.close()
    return redirect("/members")

# -------------------------
# Excel Download Routes
# -------------------------
def _time_12hr(value):
    """Convert time to 12hr string for Excel."""
    if not value:
        return '-'
    if isinstance(value, time):
        return value.strftime('%I:%M:%S %p')
    try:
        t = datetime.strptime(str(value), '%H:%M:%S')
        return t.strftime('%I:%M:%S %p')
    except ValueError:
        return str(value)

def _style_excel(ws, headers):
    """Apply styling to an Excel worksheet."""
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

@app.route('/download/attendance')
def download_attendance():
    """Download attendance as Excel with each date on a separate sheet."""
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT DISTINCT date FROM attendance ORDER BY date DESC')
    dates = [row['date'] for row in cursor.fetchall()]

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    headers = ['Name', 'Roll Number', 'Entry #', 'In Time', 'Out Time']

    for d in dates:
        date_str = d.strftime('%Y-%m-%d') if isinstance(d, date) else str(d)
        sheet_name = date_str[:31]  # Excel sheet name max 31 chars
        ws = wb.create_sheet(title=sheet_name)

        cursor.execute('''
            SELECT students.name, students.roll_number,
                   attendance.entry_number, attendance.in_time, attendance.out_time
            FROM attendance
            JOIN students ON students.id = attendance.student_id
            WHERE attendance.date = %s
            ORDER BY attendance.id
        ''', (d,))
        records = cursor.fetchall()

        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        for row_num, r in enumerate(records, 2):
            ws.cell(row=row_num, column=1, value=r['name'])
            ws.cell(row=row_num, column=2, value=r['roll_number'] or '-')
            ws.cell(row=row_num, column=3, value=r['entry_number'])
            ws.cell(row=row_num, column=4, value=_time_12hr(r['in_time']))
            ws.cell(row=row_num, column=5, value=_time_12hr(r['out_time']))

        _style_excel(ws, headers)

    conn.close()

    if not dates:
        ws = wb.create_sheet(title='No Data')
        ws.cell(row=1, column=1, value='No attendance records found')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='attendance.xlsx'
    )

@app.route('/download/members')
def download_members():
    """Download members list as Excel."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT name, roll_number, rfid_uid FROM students ORDER BY name')
    students = cursor.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Members'

    headers = ['Name', 'Roll Number', 'RFID UID']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    for row_num, s in enumerate(students, 2):
        ws.cell(row=row_num, column=1, value=s['name'])
        ws.cell(row=row_num, column=2, value=s['roll_number'] or '-')
        ws.cell(row=row_num, column=3, value=s['rfid_uid'])

    _style_excel(ws, headers)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='members.xlsx'
    )

# Local testing
if __name__ == "__main__":
    app.run(debug=True)